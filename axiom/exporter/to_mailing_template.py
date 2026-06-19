"""Конвертер источников (личная база / агентства 2GIS) в формат шаблона рассылки (Contez).

Автоопределяет тип файла, нормализует телефоны (+7XXXXXXXXXX), вытаскивает
@username и активные ссылки wa.me/t.me. Выдаёт два файла:
  * <имя>_contez.csv  — для импорта в сервис (UTF-8 с BOM, разделитель ;)
  * <имя>_contez.xlsx — то же + АКТИВНЫЕ гиперссылки WhatsApp/Telegram (кликабельны в Excel)

Запуск:
    python -m exporter.to_mailing_template data/crm100.csv
    python -m exporter.to_mailing_template data/agentstva.csv
"""
from __future__ import annotations

import csv
import io
import re
import sys
from pathlib import Path

# Колонки шаблона Contez (порядок важен).
TEMPLATE = ["Имя", "Фамилия", "Username", "Телефон", "TelegramID", "Компания",
            "Город", "Email", "Переменная1", "Переменная2", "Переменная3",
            "Переменная4", "Переменная5", "Заметки"]


# ---------- утилиты ----------

def read_text(path: str) -> str:
    raw = Path(path).read_bytes()
    for enc in ("utf-8-sig", "cp1251"):
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def norm_phone(raw: str) -> str:
    """К виду +7XXXXXXXXXX. Берёт номер до первой скобки, чистит мусор."""
    if not raw:
        return ""
    head = raw.split("(")[0]
    d = re.sub(r"\D", "", head)
    if len(d) >= 11 and d[0] in "78":
        return "+7" + d[1:11]
    if len(d) == 10:
        return "+7" + d
    return ""


def phone_from_link(link: str) -> str:
    m = re.search(r"(?:wa\.me/|number=|t\.me/\+)(\d{10,15})", link or "")
    return norm_phone(m.group(1)) if m else ""


def tg_username(*links: str) -> str:
    for link in links:
        m = re.search(r"t\.me/([A-Za-z][\w]{3,})", link or "")
        if m and not m.group(1).endswith("_bot"):
            return m.group(1)
    return ""


def first_link(*vals: str, kind: str) -> str:
    """Возвращает первую ссылку нужного типа (wa.me / t.me) как есть — для гиперссылки."""
    needle = "wa.me" if kind == "wa" else "t.me"
    for v in vals:
        if v and needle in v:
            return v.strip()
    return ""


def blank_record() -> dict:
    return {k: "" for k in TEMPLATE}


# ---------- источник 1: личная база (CRM_100, разделитель запятая) ----------

def convert_crm(text: str) -> list[dict]:
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for row in reader:
        g = {(k or "").strip().lower(): (v or "").strip() for k, v in row.items()}
        name = g.get("имя", "")
        phone_field = g.get("телефон", "")
        rec = blank_record()
        if phone_field.startswith("@") or re.fullmatch(r"[A-Za-z][\w]{3,}", phone_field):
            rec["Username"] = phone_field.lstrip("@")
        else:
            rec["Телефон"] = norm_phone(phone_field)
        rec["Имя"] = name
        rec["Город"] = g.get("город", "")
        rec["Переменная1"] = g.get("канал", "")           # макс/ватсап/нет макса
        rec["Переменная2"] = g.get("статус переговоров", "")
        notes = " | ".join(p for p in [g.get("комментарий", ""), g.get("описание задачи", "")] if p)
        rec["Заметки"] = notes
        if rec["Телефон"] or rec["Username"] or rec["Имя"]:
            out.append(rec)
    return out


# ---------- источник 2: агентства 2GIS (разделитель ;) ----------

AG = {  # индексы колонок выгрузки агентств
    "name": 0, "descr": 1, "address": 2, "city": 6,
    "tel1": 12, "tel2": 13, "email": 15, "web": 16,
    "vk": 20, "wa1": 21, "wa2": 22, "tg1": 25, "tg2": 26,
}


def convert_agency(text: str) -> list[dict]:
    reader = csv.reader(io.StringIO(text), delimiter=";")
    rows = list(reader)
    out = []
    for row in rows[1:]:
        if len(row) <= AG["tg1"] or not row[AG["name"]].strip():
            continue
        def c(key):
            return row[AG[key]].strip()
        wa_link = first_link(c("wa1"), c("wa2"), kind="wa")
        tg_link = first_link(c("tg1"), c("tg2"), kind="tg")
        phone = phone_from_link(wa_link) or norm_phone(c("tel1")) or norm_phone(c("tel2"))
        rec = blank_record()
        rec["Имя"] = c("name")
        rec["Компания"] = c("name")
        rec["Телефон"] = phone
        rec["Username"] = tg_username(c("tg1"), c("tg2"))
        rec["Город"] = c("city")
        rec["Email"] = c("email")
        rec["Переменная1"] = wa_link            # активная ссылка WhatsApp
        rec["Переменная2"] = tg_link            # активная ссылка Telegram
        rec["Переменная3"] = c("web")
        rec["Переменная4"] = c("vk")
        rec["Заметки"] = " | ".join(p for p in [c("descr"), c("address")] if p)
        if rec["Телефон"] or rec["Username"] or wa_link or tg_link:
            out.append(rec)
    return out


# ---------- запись ----------

def write_csv(records: list[dict], path: Path) -> None:
    with open(path, "w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=TEMPLATE, delimiter=";")
        w.writeheader()
        w.writerows(records)


def write_xlsx(records: list[dict], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    wb = Workbook()
    ws = wb.active
    ws.title = "Импорт"
    ws.append(TEMPLATE)
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="305496")
        cell.font = Font(color="FFFFFF", bold=True)
    link_font = Font(color="0563C1", underline="single")
    for rec in records:
        ws.append([rec[k] for k in TEMPLATE])
        r = ws.max_row
        for col_idx, key in enumerate(TEMPLATE, start=1):
            val = rec[key]
            if isinstance(val, str) and val.startswith("http"):
                cell = ws.cell(row=r, column=col_idx)
                cell.hyperlink = val      # АКТИВНАЯ ссылка
                cell.font = link_font
    for col, width in zip("ABCDEFGHIJKLMN", (20, 14, 16, 16, 12, 24, 12, 22, 26, 26, 26, 14, 12, 40)):
        ws.column_dimensions[col].width = width
    wb.save(path)


def detect_and_convert(text: str) -> tuple[str, list[dict]]:
    head = text[:200].lower()
    if ";" in head and ("наименование" in head or "whatsapp" in head):
        return "агентства", convert_agency(text)
    return "личная база", convert_crm(text)


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Укажи файл: python -m exporter.to_mailing_template data/имя.csv")
        sys.exit(1)
    src = Path(sys.argv[1])
    text = read_text(str(src))
    kind, records = detect_and_convert(text)
    base = src.with_suffix("")
    out_csv = Path(f"{base}_contez.csv")
    out_xlsx = Path(f"{base}_contez.xlsx")
    write_csv(records, out_csv)
    write_xlsx(records, out_xlsx)
    print(f"Тип: {kind}. Контактов: {len(records)}")
    print(f"CSV (для импорта):       {out_csv}")
    print(f"XLSX (активные ссылки):  {out_xlsx}")
