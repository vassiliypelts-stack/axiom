"""Импорт твоей базы контактов в книжку. Поддерживает CSV и Excel (.xlsx).

Заголовки колонок распознаются по-русски и по-английски (см. ALIASES).
Минимум — телефон ИЛИ username. Остальное опционально.

Запуск:
    python -m importer.import_contacts data/my_base.xlsx
    python -m importer.import_contacts data/contacts_example.csv
"""
from __future__ import annotations

import csv
import sys
from pathlib import Path

from db import database

# Какой заголовок в файле -> в какое поле книжки. Регистр и пробелы игнорируются.
ALIASES = {
    "phone": "phone", "телефон": "phone", "тел": "phone", "номер": "phone", "тел.": "phone",
    "username": "username", "юзернейм": "username", "telegram": "username", "тг": "username", "ник": "username",
    "name": "name", "имя": "name", "фио": "name", "контакт": "name", "имя контакта": "name",
    "city": "city", "город": "city",
    "agency": "agency", "агентство": "agency", "компания": "agency", "фирма": "agency",
    "tags": "tags", "теги": "tags", "тег": "tags", "сегмент": "tags",
    "notes": "notes", "заметки": "notes", "комментарий": "notes", "примечание": "notes", "коммент": "notes",
    # Чем человек занимается. Ради этого поля живёт главный ход холодного захода:
    # «вы тоже по этой теме? — тогда мы коллеги, поэтому и написал». Агент читает
    # его из карточки (см. channels/telegram._contact_dict), но раньше колонка с
    # описанием при импорте просто не распознавалась и текст терялся.
    "specialization": "specialization", "специализация": "specialization",
    "описание": "specialization", "description": "specialization",
    "чем занимается": "specialization", "деятельность": "specialization",
    "направление": "specialization", "о себе": "specialization",
    "about": "specialization", "bio": "specialization", "био": "specialization",
    "сфера": "specialization", "услуги": "specialization", "тема": "specialization",
    # Должность и почта — тоже уходили в никуда.
    "person_role": "person_role", "должность": "person_role", "роль": "person_role",
    "position": "person_role",
    "email": "email", "почта": "email", "e-mail": "email", "мейл": "email",
}


def sniff_delimiter(sample: str) -> str:
    """Разделитель CSV по первой строке. Русский Excel сохраняет «CSV (разделители —
    запятые)» ТОЧКОЙ С ЗАПЯТОЙ, и жёстко зашитая запятая давала «импортировано 0»
    без единой ошибки: вся строка читалась как одна колонка «Имя;Телефон;Telegram»."""
    head = (sample.splitlines() or [""])[0]
    return max((";", ",", "\t"), key=head.count) if head else ","


def normalize_phone(raw: str) -> str | None:
    """Телефон к единому виду +7XXXXXXXXXX — тому же, что у импорта 2ГИС и парсеров.

    Единый формат обязателен: дедуп контактов идёт СРАВНЕНИЕМ СТРОК (upsert_contact
    ищет по phone), поэтому «8 913…» из одного файла и «+7 913…» из другого заводили
    на одного человека две карточки — и оба получали первое сообщение."""
    if not raw:
        return None
    from importer.import_2gis import norm_phone
    s = str(raw).strip()
    p = norm_phone(s)
    if p:
        return p
    # не российский номер (или явно кривой) — оставляем как есть, только чистим мусор
    digits = "".join(ch for ch in s if ch.isdigit() or ch == "+")
    return digits or None


def normalize_username(raw: str) -> str | None:
    if not raw:
        return None
    return str(raw).strip().lstrip("@") or None


def _map_headers(headers: list[str]) -> dict[int, str]:
    """Индекс колонки -> каноничное поле."""
    out = {}
    for i, h in enumerate(headers):
        key = (h or "").strip().lower()
        if key in ALIASES:
            out[i] = ALIASES[key]
    return out


def _rows_from_csv(path: str):
    with open(path, encoding="utf-8-sig", newline="") as f:
        sample = f.readline()
        f.seek(0)
        reader = csv.reader(f, delimiter=sniff_delimiter(sample))
        headers = next(reader, [])
        colmap = _map_headers(headers)
        for row in reader:
            yield {field: (row[i].strip() if i < len(row) and row[i] else "") for i, field in colmap.items()}


def _rows_from_xlsx(path: str):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(c) if c is not None else "" for c in next(rows, [])]
    colmap = _map_headers(headers)
    for row in rows:
        if not any(row):
            continue
        yield {field: (str(row[i]).strip() if i < len(row) and row[i] is not None else "")
               for i, field in colmap.items()}


def import_file(path: str) -> int:
    database.init_db()
    ext = Path(path).suffix.lower()
    rows = _rows_from_xlsx(path) if ext in (".xlsx", ".xlsm") else _rows_from_csv(path)
    added = 0
    with database.get_conn() as conn:
        for r in rows:
            phone = normalize_phone(r.get("phone", ""))
            username = normalize_username(r.get("username", ""))
            if not phone and not username:
                continue  # пустая строка
            database.upsert_contact(
                conn,
                source="import",
                phone=phone,
                username=username,
                name=r.get("name") or None,
                city=r.get("city") or None,
                agency=r.get("agency") or None,
                tags=r.get("tags") or None,
                notes=r.get("notes") or None,
                specialization=r.get("specialization") or None,
                person_role=r.get("person_role") or None,
                email=r.get("email") or None,
            )
            added += 1
    return added


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else "data/contacts_example.csv"
    n = import_file(src)
    print(f"Импортировано/обновлено: {n} из {src}")
