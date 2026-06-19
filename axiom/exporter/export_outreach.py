"""Выгрузка контактов из книжки в Excel для сервиса рассылки (WaCombo и т.п.).

В файле и колонки-переменные (Имя/Город/Агентство — если сервис подставляет сам),
и готовая колонка «Сообщение» (если шлёшь как есть). Телефон в формате 79XXXXXXXXXX
(без +) — так его принимают WhatsApp-рассыльщики.

Запуск:
    python -m exporter.export_outreach                 # канал whatsapp (по умолчанию)
    python -m exporter.export_outreach --channel telegram
    python -m exporter.export_outreach --limit 50
"""
from __future__ import annotations

import argparse
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config
from db import database

# ⚠️ ПРАВЬ ПОД СЕБЯ. Плейсхолдеры: {name} {city} {agency}.
# Пиши так, как написал бы коллеге сам. Цель сообщения — договориться о Zoom.
OUTREACH_TEMPLATE = (
    "Здравствуйте! Пишу по агентству «{agency}»{city_part}. "
    "Я собрал систему, которая на автопилоте находит риелтору клиентов в Telegram и WhatsApp "
    "и доводит до встречи — без бюджета на рекламу. "
    "Хочу за 15 минут на Zoom показать, как это работает в недвижимости, без обязательств. "
    "Когда удобно созвониться?"
)

HEADER_FILL = PatternFill("solid", fgColor="305496")


def build_message(row) -> str:
    city = (row["city"] or "").strip()
    agency = (row["agency"] or row["name"] or "").strip()
    return OUTREACH_TEMPLATE.format(
        name=(row["name"] or "").strip(),
        agency=agency,
        city_part=f" из {city}" if city else "",
    ).strip()


def to_digits(phone: str | None) -> str:
    """79XXXXXXXXXX — формат для WhatsApp-сервисов (без +)."""
    if not phone:
        return ""
    d = re.sub(r"\D", "", phone)
    if len(d) == 11 and d[0] == "8":
        d = "7" + d[1:]
    return d


def export(channel: str = "whatsapp", limit: int | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Рассылка"
    headers = ["Телефон", "Username", "Имя", "Город", "Агентство", "Категория", "Сообщение"]
    ws.append(headers)
    for c in ws[1]:
        c.fill = HEADER_FILL
        c.font = Font(color="FFFFFF", bold=True)

    has_col = {"whatsapp": "has_wa", "telegram": "has_tg"}.get(channel)
    where = "status NOT IN ('stop','won','lost')"
    if has_col:
        where += f" AND {has_col} = 'yes'"
    q = f"SELECT name, city, agency, tags, phone, wa_phone, username FROM contacts WHERE {where} ORDER BY agency"
    if limit:
        q += f" LIMIT {int(limit)}"

    n = 0
    with database.get_conn() as conn:
        for row in conn.execute(q):
            phone = to_digits(row["wa_phone"] or row["phone"]) if channel == "whatsapp" else to_digits(row["phone"])
            ws.append([
                phone,
                (row["username"] or ""),
                (row["name"] or ""),
                (row["city"] or ""),
                (row["agency"] or ""),
                (row["tags"] or ""),
                build_message(row),
            ])
            n += 1

    for col, width in zip("ABCDEFG", (15, 18, 22, 12, 24, 22, 80)):
        ws.column_dimensions[col].width = width
    ws.column_dimensions["G"].width = 80
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=7).alignment = Alignment(wrap_text=True, vertical="top")

    out = config.BASE_DIR / f"axiom_outreach_{channel}.xlsx"
    wb.save(out)
    print(f"Выгружено {n} контактов (канал: {channel}) -> {out}")
    return str(out)


if __name__ == "__main__":
    p = argparse.ArgumentParser()
    p.add_argument("--channel", default="whatsapp", choices=["whatsapp", "telegram", "all"])
    p.add_argument("--limit", type=int, default=None)
    args = p.parse_args()
    export(args.channel, args.limit)
