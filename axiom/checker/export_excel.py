"""Выгрузка результатов чекера в Excel с подсветкой:
зелёный = есть мессенджер, красный = нет, серый = не проверено.

Запуск:  python -m checker.export_excel
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

import config
from db import database

GREEN = PatternFill("solid", fgColor="C6EFCE")   # есть
RED = PatternFill("solid", fgColor="FFC7CE")      # нет
GREY = PatternFill("solid", fgColor="D9D9D9")     # не проверено
HEADER = PatternFill("solid", fgColor="305496")

FILLS = {"yes": GREEN, "no": RED, "unknown": GREY}
LABEL = {"yes": "есть", "no": "нет", "unknown": "?"}


def export(path: str = "axiom_checker.xlsx") -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "Чекер"

    headers = ["Имя", "Телефон", "Город", "Telegram", "WhatsApp", "MAX"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    with database.get_conn() as conn:
        rows = conn.execute(
            "SELECT name, phone, city, has_tg, has_wa, has_max FROM contacts ORDER BY name"
        ).fetchall()

    for r in rows:
        ws.append([r["name"], r["phone"], r["city"],
                   LABEL.get(r["has_tg"], "?"), LABEL.get(r["has_wa"], "?"), LABEL.get(r["has_max"], "?")])
        row_idx = ws.max_row
        for col, key in zip((4, 5, 6), (r["has_tg"], r["has_wa"], r["has_max"])):
            c = ws.cell(row=row_idx, column=col)
            c.fill = FILLS.get(key, GREY)
            c.alignment = Alignment(horizontal="center")

    for col, width in zip("ABCDEF", (22, 16, 14, 11, 11, 8)):
        ws.column_dimensions[col].width = width

    out = config.BASE_DIR / path
    wb.save(out)
    return str(out)


if __name__ == "__main__":
    print(f"Сохранено: {export()}")
